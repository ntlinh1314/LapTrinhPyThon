from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox
from openpyxl import *
import re

wb = load_workbook('E:\Lưu bài Python\Lab4_2110060_NguyenThuyLinh\Bai6.xlsx')
sheet = wb.active
 
def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
 
    sheet.cell(row=1, column=1).value = "Ma so sinh vien"
    sheet.cell(row=1, column=2).value = "Ho ten"
    sheet.cell(row=1, column=3).value = "Ngay sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "So dien thoai"
    sheet.cell(row=1, column=6).value = "Hoc ky"
    sheet.cell(row=1, column=7).value = "Nam hoc"
    sheet.cell(row=1, column=8).value = "Mon hoc"
 
def clear():
    mssv.delete(0, END)
    hoten.delete(0, END)
    ngaysinh.delete(0, END)
    email.delete(0, END)
    sdt.delete(0, END)
    hocky.delete(0, END)
    namhoc.delete(0, END)

def insert():
    if (mssv.get() == "" and
        hoten.get() == "" and
        ngaysinh.get() == "" and
        email.get() == "" and
        sdt.get() == "" and
        hocky.get() == "" and
        namhoc.get() == ""):
        print("empty input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row=current_row + 1, column=1).value = mssv.get()
        sheet.cell(row=current_row + 1, column=2).value = hoten.get()
        sheet.cell(row=current_row + 1, column=3).value = ngaysinh.get()
        sheet.cell(row=current_row + 1, column=4).value = email.get()
        sheet.cell(row=current_row + 1, column=5).value = sdt.get()
        sheet.cell(row=current_row + 1, column=6).value = hocky.get()
        sheet.cell(row=current_row + 1, column=7).value = namhoc.get()
        sheet.cell(row=current_row + 1, column=8).value = chonMon1.get()

        wb.save('E:\Lưu bài Python\Lab4_2110060_NguyenThuyLinh\Bai6.xlsx')

        mssv.focus_set()
 
        clear()

special_ch = ['~', '`', '!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '-', '_', '+', '=', '{', '}', '[', ']', '|', '\\', '/', ':', ';', '"', "'", '<', '>', ',', '.', '?']

def KtraMS():
    maSo = mssv.get()
    msg = ""
    if len(maSo) == 0:
        msg = 'Ma so khong duoc de trong!'
    else:
        try:
            if any(tu in special_ch for tu in maSo):
                msg = 'Khong duoc phep nhap ky tu'
            elif any(tu.isupper() or tu.islower() for tu in maSo):
                msg = 'Khong duoc phep nhap chu'
            elif len(maSo) < 7:
                msg = 'Ma so phai co do dai bang 7!'
        except Exception as ep:
            messagebox.showerror('error', ep)
    messagebox.showinfo('Thong Bao', msg)  

def KtraMail():
    mail = email.get()
    regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    if not re.match(regex, mail):
        email.delete(0, END)
        messagebox.showerror(0, "Loi dinh dang Email")

def KtraSDT():
    soDT = sdt.get()
    msg = ""
    if len(soDT) == 0:
        msg = 'So dien thoai khong duoc de trong!'
    else:
        try:
            if len(soDT) < 10:
                msg = 'So dien thoai phai co do dai bang 10!'
        except Exception as ep:
            messagebox.showerror('error', ep)
    messagebox.showinfo('Thong Bao', msg)

def KtraHocKy():
    hocKy = hocky.get()
    if hocKy not in ['1','2','3']:
        hocky.delete(0, END)
        messagebox.showerror(0,"Chi duoc nhap vao hoc ky (1, 2, 3)")

def KtraNgay():
    ngay = ngaysinh.get()
    dk = r'^\d{2}/\d{2}/\d{4}$'
    if not re.match(dk, ngay):
        ngaysinh.delete(0, END)
        messagebox.showerror(0, "Dinh dang ngay sai (dd/mm/yyyy)")

def Thoat():
    root.destroy()

if __name__ == "__main__":
    root = Tk()
    root.title("Dang Ky Hoc Phan")
    root.geometry("500x500")
    root.config(bg="#98ec94")
    
    heading = Label(root, text="THONG TIN DANG KY HOC PHAN", fg="red", bg="#98ec94",font=30)
    mssv_lb = Label(root, text="Ma so sinh vien",bg="#98ec94")
    hoten_lb = Label(root, text="Ho ten",bg="#98ec94")
    ngaysinh_lb = Label(root, text="Ngay sinh",bg="#98ec94")
    email_lb = Label(root, text="Email",bg="#98ec94")
    sdt_lb = Label(root, text="So dien thoai",bg="#98ec94")
    hocky_lb = Label(root, text="Hoc ky",bg="#98ec94")
    namhoc_lb = Label(root, text="Nam hoc",bg="#98ec94")
    chon_lb = Label(root, text="Chon mon hoc",bg="#98ec94")

    heading.place(anchor='center', relx=0.5, rely=0.1)
    mssv_lb.place(relx=0.01, rely=0.15)
    hoten_lb.place(relx=0.01,rely=0.2)
    ngaysinh_lb.place(relx=0.01,rely=0.25)
    email_lb.place(relx=0.01, rely=0.3)
    sdt_lb.place(relx=0.01, rely=0.35)
    hocky_lb.place(relx=0.01, rely=0.4)
    namhoc_lb.place(relx=0.01,rely=0.45)
    chon_lb.place(relx=0.01, rely=0.5)

    mssv = Entry(root, width=60)
    hoten = Entry(root, width=60)
    ngaysinh = Entry(root,width=60)
    email = Entry(root, width=60)
    sdt = Entry(root, width=60)
    hocky = Entry(root, width=60)
    n = StringVar()
    namhoc = Combobox(root, width=57, textvariable= n)
    namhoc['values']=('2022-2023','2023-2024','2024-2025')
    namhoc.current(1)

    excel()

    chonMon1 = StringVar()
    chonMon2= StringVar()
    chonMon3 = StringVar()
    chonMon4 = StringVar()
    chonMon_check = Checkbutton(root,text="Lap trinh Python",bg="#98ec94",variable=chonMon1).place(relx=0.24, rely=0.5)
    chonMon_check = Checkbutton(root,text="Cong nghe phan mem",bg="#98ec94",variable=chonMon2).place(relx=0.24, rely=0.55)
    chonMon_check = Checkbutton(root,text="Lap trinh Java",bg="#98ec94", variable=chonMon3).place(relx=0.64, rely=0.5)
    chonMon_check = Checkbutton(root,text="Phat trien ung dung web",bg="#98ec94",variable=chonMon4).place(relx=0.64, rely=0.55)
    Button(root, text="Dang ky",bg="lime",command=lambda: [KtraMS(), KtraMail(), KtraHocKy(), KtraNgay(),KtraSDT(), insert()]).place(relx=0.35, rely=0.65)
    Button(root, text="Thoat",bg="lime",command=Thoat).place(relx=0.65, rely=0.65)
    

    mssv.place(relx=0.25, rely=0.15)
    hoten.place(relx=0.25, rely=0.2)
    ngaysinh.place(relx=0.25, rely=0.25)
    email.place(relx=0.25, rely=0.3)
    sdt.place(relx=0.25, rely=0.35)
    hocky.place(relx=0.25, rely=0.4)
    namhoc.place(relx=0.25, rely=0.45)
    root.mainloop()